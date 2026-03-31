#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel to TIA Portal SCL Code Generator

Программа для парсинга таблиц Excel с разметкой системы ввода-вывода
и генерации готового кода для TIA Portal (SCL).
"""

import openpyxl
import re
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple
from pathlib import Path
from datetime import datetime


@dataclass
class DeviceConfig:
    """Конфигурация устройства"""
    place: int  # Номер ванны/позиции
    device_num: int  # Номер устройства
    device_type: int  # Тип устройства
    aux_place: Optional[int] = None  # Вспомогательная позиция
    aux_device: Optional[int] = None  # Вспомогательное устройство
    comment: str = ""  # Комментарий
    extra_params: Dict[str, any] = field(default_factory=dict)  # Дополнительные параметры


@dataclass
class ModuleChannel:
    """Канал модуля ввода/вывода"""
    module_name: str  # Имя модуля (А14, А24 и т.д.)
    module_addr: int  # Адрес модуля
    channel_num: int  # Номер канала
    signal_name: str  # Имя сигнала
    place: int  # Номер ванны
    device_type: str  # Тип устройства
    cabinet: str = ""  # Шкаф
    cabinet_type: str = ""  # Тип шкафа (DO, DI, AI, AO)


@dataclass
class GeneratedCode:
    """Сгенерированный код"""
    region_name: str
    code_lines: List[str]
    description: str = ""


class ExcelToSCLConverter:
    """Конвертер Excel в SCL код"""
    
    # Маппинг типов оборудования из Excel в имена регионов SCL
    EQUIPMENT_TYPE_MAP = {
        'Температура': 'Temperature',
        'Нагрев': 'Temperature',
        'Долив': 'Doliv',
        'Жироуловитель': 'Jr',
        'Перемешивание': 'Mixer',
        'Барботаж': 'Mixer',
        'Выпрямитель': 'Vip',
        'Фильтрование': 'Filtr',
        'Дозирование': 'Doser',
        'Душирование': 'Shower',
        'Качание': 'Pok',
        'Сушка': 'Dry',
        'Слив': 'Sink',
        'Крышки': 'Cover',
        'Воздуходувка': 'Blower',
        'Чиллер': 'Chiller',
        'Барьер безопасности': 'SafetyBar',
        'Вентиляция': 'VentAbsorb',
        'Перемещение': 'Cart',
        'Вращение барабанов': 'LineDev',
    }
    
    # Параметры по умолчанию для разных типов устройств
    DEFAULT_PARAMS = {
        'Doliv': {'CfgAutoLevel': True},
        'Mixer': {'CfgPnevmo': True},
        'Cover': {'CfgPnevmo': True},
        'Jr': {'CfgAutoLevel': True},
    }
    
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.wb = openpyxl.load_workbook(excel_path, data_only=True)
        self.devices: Dict[str, List[DeviceConfig]] = {}
        self.module_channels: Dict[str, List[ModuleChannel]] = {
            'AI': [],
            'AO': [],
            'DI': [],
            'DO': []
        }
        self.max_counts: Dict[str, int] = {}
        
    def resolve_formula(self, value, row_cache: dict) -> any:
        """Разрешение формул Excel вида =$A$7"""
        if value is None:
            return None
            
        if isinstance(value, str) and value.startswith('=$'):
            # Извлекаем ссылку на ячейку
            match = re.match(r'=\$([A-Z]+)\$(\d+)', value)
            if match:
                col, row = match.groups()
                cache_key = f"{col}{row}"
                if cache_key in row_cache:
                    return row_cache[cache_key]
                return value
        return value
    
    def parse_sheet_shs(self):
        """Парсинг листа ШС (шкаф силовой)"""
        ws = self.wb['ШС']
        rows = list(ws.iter_rows(values_only=True))
        
        # Кэш для разрешения формул
        row_cache = {}
        
        # Находим заголовки (строка 5-6)
        # Строка 5: Устройство ввода/вывода, Device, Наименование сигнала...
        # Строка 6: шкаф, № п/п, обозн., тип, адрес, № вх....
        
        for row_idx in range(6, len(rows)):  # Начинаем с 7-й строки (индекс 6)
            row = rows[row_idx]
            if not row or all(cell is None for cell in row[:6]):
                continue
                
            # Разрешаем формулы
            resolved_row = []
            for col_idx, cell in enumerate(row):
                col_letter = chr(ord('A') + col_idx)
                cache_key = f"{col_letter}{row_idx + 1}"
                
                if isinstance(cell, str) and cell.startswith('=$'):
                    match = re.match(r'=\$([A-Z]+)\$(\d+)', cell)
                    if match:
                        ref_col, ref_row = match.groups()
                        ref_row_idx = int(ref_row) - 1
                        if ref_row_idx < len(rows):
                            ref_cell = rows[ref_row_idx][col_idx]
                            resolved_row.append(ref_cell if ref_cell else cell)
                            row_cache[cache_key] = ref_cell if ref_cell else cell
                            continue
                
                resolved_row.append(cell)
                if cell and not (isinstance(cell, str) and cell.startswith('=$')):
                    row_cache[cache_key] = cell
            
            row = resolved_row
            
            cabinet = row[0]  # Шкаф (ШС1)
            cabinet_no = row[1]  # № п/п
            designation = row[2]  # Обозначение (А14)
            cabinet_type = row[3]  # Тип (DO, AO и т.д.)
            address = row[4]  # Адрес
            channel_num = row[5]  # № вх.
            signal_name = row[6]  # Наименование сигнала
            tech_pos = row[7]  # Тех. поз.
            main_pos = row[8]  # № п/п основ. поз.
            aux_pos = row[9]  # № п/п вспомогат. поз.
            comment = row[15] if len(row) > 15 else None  # Примечание
            
            if not designation or not channel_num:
                continue
            
            # Определяем тип модуля
            if cabinet_type:
                cabinet_type = cabinet_type.strip()
            
            # Создаем запись канала модуля
            channel = ModuleChannel(
                module_name=designation,
                module_addr=address if address else 0,
                channel_num=channel_num,
                signal_name=signal_name or "",
                place=self._parse_place(main_pos),
                device_type=signal_name or "",
                cabinet=cabinet or "",
                cabinet_type=cabinet_type or ""
            )
            
            # Добавляем в соответствующий список
            if cabinet_type:
                if 'DO' in cabinet_type:
                    self.module_channels['DO'].append(channel)
                elif 'AO' in cabinet_type:
                    self.module_channels['AO'].append(channel)
                elif 'DI' in cabinet_type:
                    self.module_channels['DI'].append(channel)
                elif 'AI' in cabinet_type:
                    self.module_channels['AI'].append(channel)
            
            # Парсим конфигурацию устройства
            self._parse_device_config(signal_name, main_pos, aux_pos, comment)
    
    def parse_sheet_shsau(self):
        """Парсинг листа ШСАУ (шкаф автоматики)"""
        ws = self.wb['ШСАУ']
        rows = list(ws.iter_rows(values_only=True))
        
        row_cache = {}
        
        for row_idx in range(2, len(rows)):  # Начинаем с 3-й строки (индекс 2)
            row = rows[row_idx]
            if not row or all(cell is None for cell in row[:6]):
                continue
            
            # Разрешаем формулы
            resolved_row = []
            for col_idx, cell in enumerate(row):
                col_letter = chr(ord('A') + col_idx)
                cache_key = f"{col_letter}{row_idx + 1}"
                
                if isinstance(cell, str) and cell.startswith('=$'):
                    match = re.match(r'=\$([A-Z]+)\$(\d+)', cell)
                    if match:
                        ref_col, ref_row = match.groups()
                        ref_row_idx = int(ref_row) - 1
                        if ref_row_idx < len(rows):
                            ref_cell = rows[ref_row_idx][col_idx]
                            resolved_row.append(ref_cell if ref_cell else cell)
                            row_cache[cache_key] = ref_cell if ref_cell else cell
                            continue
                
                resolved_row.append(cell)
                if cell and not (isinstance(cell, str) and cell.startswith('=$')):
                    row_cache[cache_key] = cell
            
            row = resolved_row
            
            cabinet = row[0]  # Шкаф (ШСАУ1)
            cabinet_no = row[1]  # № п/п
            designation = row[2]  # Обозначение
            cabinet_type = row[3]  # Тип (DI, DO и т.д.)
            address = row[4]  # Адрес
            channel_num = row[5]  # № вх.
            signal_name = row[7]  # Наименование сигнала
            tech_pos = row[8]  # Тех. поз.
            main_pos = row[9]  # № п/п основ. поз.
            comment = row[15] if len(row) > 15 else None  # Примечание
            
            if not designation or not channel_num:
                continue
            
            if cabinet_type:
                cabinet_type = cabinet_type.strip()
            
            channel = ModuleChannel(
                module_name=designation,
                module_addr=address if address else 0,
                channel_num=channel_num,
                signal_name=signal_name or "",
                place=self._parse_place(main_pos),
                device_type=signal_name or "",
                cabinet=cabinet or "",
                cabinet_type=cabinet_type or ""
            )
            
            if cabinet_type:
                if 'DI' in cabinet_type:
                    self.module_channels['DI'].append(channel)
                elif 'DO' in cabinet_type:
                    self.module_channels['DO'].append(channel)
                elif 'AI' in cabinet_type:
                    self.module_channels['AI'].append(channel)
                elif 'AO' in cabinet_type:
                    self.module_channels['AO'].append(channel)
    
    def _parse_place(self, place_value) -> int:
        """Парсинг номера позиции (может быть "11, 12" или просто число)"""
        if place_value is None:
            return 0
        if isinstance(place_value, (int, float)):
            return int(place_value)
        if isinstance(place_value, str):
            # Может быть "11, 12" или "2-10"
            match = re.search(r'(\d+)', place_value)
            if match:
                return int(match.group(1))
        return 0
    
    def _parse_device_config(self, signal_name: str, main_pos, aux_pos, comment: str = None):
        """Парсинг конфигурации устройства"""
        if not signal_name:
            return
            
        # Определяем тип региона
        region_name = None
        for key, value in self.EQUIPMENT_TYPE_MAP.items():
            if key in signal_name:
                region_name = value
                break
        
        if not region_name:
            return
        
        # Парсим номера позиций
        place = self._parse_place(main_pos)
        aux_place = self._parse_place(aux_pos)
        
        if place == 0:
            return
        
        # Определяем номер устройства (по порядку в регионе)
        if region_name not in self.devices:
            self.devices[region_name] = []
        
        # Проверяем, не добавлено ли уже это устройство
        existing = [d for d in self.devices[region_name] if d.place == place]
        if existing:
            return
        
        device_num = len(self.devices[region_name]) + 1
        
        # Определяем тип устройства (можно расширить)
        device_type = 1  # По умолчанию
        
        config = DeviceConfig(
            place=place,
            device_num=device_num,
            device_type=device_type,
            aux_place=aux_place if aux_place != place else None,
            comment=comment or ""
        )
        
        self.devices[region_name].append(config)
    
    def generate_max_counts_code(self) -> str:
        """Генерация кода для установки максимальных количеств устройств"""
        lines = [
            '// Установленное количество устройств',
        ]
        
        region_order = ['OP', 'Row', 'AO', 'Cart', 'Vann', 'Doliv', 'Temperature', 
                       'Cover', 'Jr', 'Mixer', 'Vip', 'Filtr', 'Doser', 'Shower', 
                       'Pok', 'Dry', 'Sink', 'PID', 'Blower', 'Chiller', 'SafetyBar', 'Lifter']
        
        for region in region_order:
            count = len(self.devices.get(region, []))
            lines.append(f'    "Options".Count.Max{region} := {count};')
        
        return '\n'.join(lines) + ';'
    
    def generate_region_code(self, region_name: str) -> Optional[str]:
        """Генерация кода для конкретного региона"""
        if region_name not in self.devices or not self.devices[region_name]:
            return None
        
        devices = sorted(self.devices[region_name], key=lambda x: x.place)
        lines = []
        
        # Заголовок региона
        lines.append(f'REGION {region_name}')
        
        # Комментарии в зависимости от типа
        if region_name == 'Doliv':
            lines.extend([
                '(* Доливы',
                'CfgPlace        номер ванны',
                'CfgType         тип',
                'CfgCascadeVann  для каскадных доливов указываем номер второй ванны каскада',
                'CfgAutoLevel    автоматическое получение уровня из ванны по Place',
                '*)'
            ])
            # Инициализация по умолчанию
            lines.append(f'    #iCnt := 1;')
            lines.append(f'    WHILE #iCnt <= "Options".Count.MaxDoliv DO')
            lines.append(f'        "Doliv".Dev[#iCnt].CfgAutoLevel := TRUE;')
            lines.append(f'        #iCnt := #iCnt + 1;')
            lines.append(f'    END_WHILE;')
            lines.append('')
            
            # Конфигурация устройств
            for dev in devices:
                lines.append(f'    "Doliv".Dev[{dev.device_num}].CfgPlace := {dev.place};')
                lines.append(f'    "Doliv".Dev[{dev.device_num}].CfgType := {dev.device_type};')
                if dev.aux_place and dev.aux_place != dev.place:
                    lines.append(f'    "Doliv".Dev[{dev.device_num}].CfgCascadeVann := {dev.aux_place};')
                lines.append('')
        
        elif region_name == 'Temperature':
            lines.extend([
                '(* Нагревы и охлаждения',
                'CfgPlace  номер ванны',
                'CfgType   тип',
                '*)'
            ])
            lines.append('')
            for dev in devices:
                lines.append(f'    "Tmpr".Dev[{dev.device_num}].CfgPlace := {dev.place};')
                lines.append(f'    "Tmpr".Dev[{dev.device_num}].CfgType := {dev.device_type};')
                lines.append('')
        
        elif region_name == 'Mixer':
            lines.extend([
                '(* Мешалки, барботажи',
                'CfgPlace  номер ванны',
                'CfgType   тип',
                'CfgPnevmo по умолчанию перемешивание воздухом',
                '*)'
            ])
            lines.append(f'    #iCnt := 1;')
            lines.append(f'    WHILE #iCnt <= "Options".Count.MaxMixer DO')
            lines.append(f'        "Mixer".Dev[#iCnt].CfgPnevmo := TRUE;')
            lines.append(f'        #iCnt := #iCnt + 1;')
            lines.append(f'    END_WHILE;')
            lines.append('')
            
            for dev in devices:
                lines.append(f'    "Mixer".Dev[{dev.device_num}].CfgPlace := {dev.place};')
                lines.append(f'    "Mixer".Dev[{dev.device_num}].CfgType := {dev.device_type};')
                lines.append('')
        
        elif region_name == 'Cover':
            lines.extend([
                '(* Крышки ванн',
                'CfgPlace  номер ванны',
                'CfgType   тип',
                'CfgPnevmo крышки работают на пневматике',
                '*)'
            ])
            lines.append(f'    #iCnt := 1;')
            lines.append(f'    WHILE #iCnt <= "Options".Count.MaxCover DO')
            lines.append(f'        "Cover".Dev[#iCnt].CfgPnevmo := TRUE;')
            lines.append(f'        #iCnt := #iCnt + 1;')
            lines.append(f'    END_WHILE;')
            lines.append('')
            
            for dev in devices:
                lines.append(f'    "Cover".Dev[{dev.device_num}].CfgPlace := {dev.place};')
                lines.append(f'    "Cover".Dev[{dev.device_num}].CfgType := {dev.device_type};')
                lines.append('')
        
        elif region_name == 'Jr':
            lines.extend([
                '(* Жироуловители',
                'CfgPlace     номер ванны',
                'CfgType      тип',
                'CfgAutoLevel автоматическое получение уровня',
                '*)'
            ])
            lines.append(f'    #iCnt := 1;')
            lines.append(f'    WHILE #iCnt <= "Options".Count.MaxJr DO')
            lines.append(f'        "Jr".Dev[#iCnt].CfgAutoLevel := TRUE;')
            lines.append(f'        #iCnt := #iCnt + 1;')
            lines.append(f'    END_WHILE;')
            lines.append('')
            
            for dev in devices:
                lines.append(f'    "Jr".Dev[{dev.device_num}].CfgPlace := {dev.place};')
                lines.append(f'    "Jr".Dev[{dev.device_num}].CfgType := {dev.device_type};')
                lines.append('')
        
        elif region_name == 'Filtr':
            lines.extend([
                '(* Фильтровалки',
                'CfgPlace  номер ванны',
                'CfgType   тип',
                '*)'
            ])
            lines.append('')
            for dev in devices:
                lines.append(f'    "Filtr".Dev[{dev.device_num}].CfgPlace := {dev.place};')
                lines.append(f'    "Filtr".Dev[{dev.device_num}].CfgType := {dev.device_type};')
                lines.append('')
        
        elif region_name == 'SafetyBar':
            lines.extend([
                '// Барьеры безопасности',
            ])
            for dev in devices:
                lines.append(f'    "SafetyBar".Dev[{dev.device_num}].CfgPlace := {dev.place};')
                lines.append(f'    "SafetyBar".Dev[{dev.device_num}].CfgInRow := 0;')
                lines.append(f'    "SafetyBar".Dev[{dev.device_num}].CfgNoErrorMessage := FALSE;')
                lines.append('')
        
        lines.append(f'END_REGION')
        
        return '\n'.join(lines)
    
    def generate_module_ai_code(self) -> str:
        """Генерация кода для аналоговых входов (AI)"""
        channels = self.module_channels.get('AI', [])
        if not channels:
            channels = self.module_channels.get('AO', [])  # Пробуем AO если нет AI
        
        # Группируем по модулям
        modules: Dict[str, List[ModuleChannel]] = {}
        for ch in channels:
            if ch.module_name not in modules:
                modules[ch.module_name] = []
            modules[ch.module_name].append(ch)
        
        if not modules:
            return '// Нет аналоговых входов для генерации\n'
        
        lines = []
        for module_name, module_channels in sorted(modules.items()):
            lines.append(f'REGION {module_name}')
            
            # Находим адрес модуля
            addr = module_channels[0].module_addr if module_channels else 0
            lines.append(f'    // {module_name}. Адрес {addr}')
            
            # Сортируем каналы по номеру
            module_channels.sort(key=lambda x: x.channel_num)
            
            for ch in module_channels:
                # Определяем переменную назначения на основе типа сигнала
                target_var = self._get_ai_target_variable(ch.signal_name, ch.place)
                if target_var:
                    module_idx = list(modules.keys()).index(module_name) + 1
                    lines.append(f'    {target_var} := "MapAin".Module[{module_idx}].Channel[{ch.channel_num}];')
            
            lines.append('END_REGION\n')
        
        return '\n'.join(lines)
    
    def _get_ai_target_variable(self, signal_name: str, place: int) -> Optional[str]:
        """Определение целевой переменной для аналогового входа"""
        if 'Температура' in signal_name:
            return f'"Vann".Dev[{place}].Temperature'
        elif 'Давление' in signal_name:
            return f'"Vann".Dev[{place}].Pressure'
        elif 'pH' in signal_name:
            return f'"Vann".Dev[{place}].pH'
        elif 'Уровень' in signal_name:
            return f'"Vann".Dev[{place}].Level'
        return None
    
    def generate_module_do_code(self) -> str:
        """Генерация кода для цифровых выходов (DO)"""
        channels = self.module_channels.get('DO', [])
        
        # Группируем по модулям
        modules: Dict[str, List[ModuleChannel]] = {}
        for ch in channels:
            if ch.module_name not in modules:
                modules[ch.module_name] = []
            modules[ch.module_name].append(ch)
        
        if not modules:
            return '// Нет цифровых выходов для генерации\n'
        
        lines = []
        for module_name, module_channels in sorted(modules.items()):
            lines.append(f'REGION {module_name}')
            
            addr = module_channels[0].module_addr if module_channels else 0
            lines.append(f'    // {module_name}. Адрес {addr}')
            lines.append('    #dwModuleBitMask := 0;  // обнулим маску выходов')
            lines.append('')
            
            # Сортируем каналы по номеру
            module_channels.sort(key=lambda x: x.channel_num)
            
            # Генерируем биты (по 32 бита на модуль)
            bit_assignments = []
            for i, ch in enumerate(module_channels[:32], 1):
                bit_var = self._get_do_bit_variable(ch)
                if bit_var:
                    bit_assignments.append((i, bit_var, ch.signal_name))
            
            # Группируем по 8 битов для читаемости
            for i in range(0, len(bit_assignments), 8):
                chunk = bit_assignments[i:i+8]
                for bit_num, bit_var, signal_name in chunk:
                    lines.append(f'    #xBit.b{bit_num} := {bit_var};')
                lines.append('')
            
            lines.append('    "MapDout".Module[MODULE_IDX].BitMask := #dwModuleBitMask;')
            lines.append('')
            lines.append('END_REGION\n')
        
        return '\n'.join(lines)
    
    def _get_do_bit_variable(self, ch: ModuleChannel) -> Optional[str]:
        """Определение переменной для бита цифрового выхода"""
        signal = ch.signal_name
        
        if 'Температура' in signal or 'Нагрев' in signal:
            # Нужно определить номер устройства
            return f'"Tmpr".Dev[PLACE].qStage[1]'
        elif 'Долив' in signal:
            return f'"Doliv".Dev[PLACE].qDoliv'
        elif 'Перемешивание' in signal or 'Барботаж' in signal:
            return f'"Mixer".Dev[PLACE].qBarb'
        elif 'Крышк' in signal:
            return f'"Cover".Dev[PLACE].qOpen'
        elif 'Фильтр' in signal:
            return f'"Filtr".Dev[PLACE].qFiltr'
        elif 'Вентилятор' in signal:
            return f'"Vent".Dev[PLACE].qRun'
        elif 'Насос' in signal:
            return f'"Pump".Dev[PLACE].qRun'
        
        return None
    
    def generate_all(self, output_path: str):
        """Генерация всего кода и сохранение в файл"""
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        
        lines = [
            f'// Сгенерировано из {self.excel_path.name}',
            f'// Дата: {datetime.now().strftime("%d.%m.%Y %H:%M")}',
            '',
            self.generate_max_counts_code(),
            '',
        ]
        
        # Порядок регионов для вывода
        region_order = ['Doliv', 'Temperature', 'Cover', 'Jr', 'Mixer', 'Vip', 
                       'Filtr', 'Doser', 'Shower', 'Pok', 'Dry', 'Sink', 
                       'PID', 'Blower', 'Chiller', 'SafetyBar', 'VentAbsorb']
        
        for region in region_order:
            code = self.generate_region_code(region)
            if code:
                lines.append(code)
                lines.append('')
        
        # Модули ввода/вывода
        lines.append('// Пример для AI')
        lines.append(self.generate_module_ai_code())
        lines.append('')
        
        lines.append('// Пример для DO')
        lines.append(self.generate_module_do_code())
        
        # Записываем в файл
        with open(output, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        print(f"Код сгенерирован и сохранен в {output}")
        return str(output)


def main():
    """Основная функция"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Генератор SCL кода для TIA Portal из Excel файлов'
    )
    parser.add_argument(
        'excel_file',
        help='Путь к Excel файлу с системой ввода-вывода'
    )
    parser.add_argument(
        '-o', '--output',
        default='generated_code.txt',
        help='Путь к выходному файлу (по умолчанию: generated_code.txt)'
    )
    
    args = parser.parse_args()
    
    # Проверяем существование файла
    if not Path(args.excel_file).exists():
        print(f"Ошибка: Файл {args.excel_file} не найден")
        return 1
    
    try:
        converter = ExcelToSCLConverter(args.excel_file)
        
        # Парсим листы
        print("Парсинг листа ШС...")
        converter.parse_sheet_shs()
        
        if 'ШСАУ' in converter.wb.sheetnames:
            print("Парсинг листа ШСАУ...")
            converter.parse_sheet_shsau()
        
        # Статистика
        print(f"\nНайдено устройств по регионам:")
        for region, devices in sorted(converter.devices.items()):
            print(f"  {region}: {len(devices)} устройств")
        
        print(f"\nНайдено каналов модулей:")
        for module_type, channels in converter.module_channels.items():
            if channels:
                print(f"  {module_type}: {len(channels)} каналов")
        
        # Генерируем код
        print(f"\nГенерация кода...")
        output_file = converter.generate_all(args.output)
        
        print(f"\nГотово! Код сохранен в {output_file}")
        return 0
        
    except Exception as e:
        print(f"Ошибка: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    exit(main())
